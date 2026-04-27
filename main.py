from pathlib import Path
from datetime import datetime, date
from uuid import uuid4
from copy import copy
import json
import base64
import shutil
import io
import traceback
import subprocess
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from urllib.parse import quote_plus

from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from PIL import Image as PILImage
import sqlite3

try:
    import qrcode
except ImportError:
    qrcode = None

from database import init_db, insert_verificateur, get_all_verificateurs, search_verificateurs


# =========================================================
# CONFIG
# =========================================================

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Omnilux2026")
APP_SECRET_KEY = os.getenv("APP_SECRET_KEY", "SUPER_SECRET_KEY_CHANGE_MOI")

APP_PUBLIC_URL = os.getenv("APP_PUBLIC_URL", "http://127.0.0.1:8000").rstrip("/")

# SQLite - Phase 1 MVP
# Base de données simple dans un fichier à la racine du projet.
SQLITE_DB_PATH = Path(os.getenv("SQLITE_DB_PATH", "./echaff.db")).resolve()

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.office365.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USER)

MAX_SOCIETES_UTILISATRICES = 10

ROW_START_SOCIETES = 33
ROW_END_SOCIETES = 42

COL_SOCIETE = "A"
COL_REPRESENTANT = "H"
COL_DATE = "O"
COL_SIGNATURE = "T"

# ---- Zone vérificateur
VERIF_SIGNATURE_CELL = "AC38:AR38"
VERIF_NAME_CELL = "AC35"
VERIF_DATE_CELL = "AC37"
VERIF_HOUR_CELL = "AP37"
VERIF_DIPLOME_CELL = "AP35"

# ---- Zone entreprise utilisatrice / maître d'œuvre
CLIENT_NAME_CELL = "AI41"
CLIENT_DATE_CELL = "AC42"
CLIENT_HOUR_CELL = "AP42"
CLIENT_SIGNATURE_CELL = "AC43"

# ---- Type d'entreprise
TYPE_ENTREPRISE_TEXT_CELL = "AB33"

# ---- Observations
OBSERVATIONS_CELL = "A48"
PRINT_AREA = "$A$1:$AR$124"

ATTENTE_FILL = PatternFill(fill_type="solid", fgColor="FF4D4D")
ATTENTE_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
ATTENTE_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
SIGNATURES_DIR = BASE_DIR / "signatures"

UPLOADS_DIR = BASE_DIR / "uploads"
CARTES_DIR = UPLOADS_DIR / "cartes_identite"
DIPLOMES_DIR = UPLOADS_DIR / "diplomes"
QR_CODES_DIR = UPLOADS_DIR / "qr_codes"


# =========================================================
# INITIALISATION DOSSIERS
# =========================================================

for directory in [
    TEMPLATES_DIR,
    DATA_DIR,
    OUTPUT_DIR,
    SIGNATURES_DIR,
    UPLOADS_DIR,
    CARTES_DIR,
    DIPLOMES_DIR,
    QR_CODES_DIR,
]:
    directory.mkdir(parents=True, exist_ok=True)


# =========================================================
# APP
# =========================================================

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=APP_SECRET_KEY)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")
app.mount("/output", StaticFiles(directory=str(OUTPUT_DIR)), name="output")
app.mount("/data", StaticFiles(directory=str(DATA_DIR)), name="data")


# =========================================================
# STARTUP
# =========================================================

@app.on_event("startup")
def startup_event():
    init_db()
    init_app_db()


@app.get("/test")
def test():
    return {"status": "ok"}

# AJOUT TEMPORAIRE
@app.get("/debug/db")
def debug_db():
    import sqlite3
    from pathlib import Path

    db_path = Path("./echaff.db").resolve()

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [row[0] for row in cur.fetchall()]
    conn.close()

    return {
        "db_path": str(db_path),
        "exists": db_path.exists(),
        "tables": tables
    }

# =========================================================
# DETECTION DU MODELE EXCEL
# =========================================================

def find_excel_template() -> Path:
    candidates = [
        TEMPLATES_DIR / "PV_MODELE.xlsx",
        TEMPLATES_DIR / "PV_MODELE.xlsm",
        TEMPLATES_DIR / "PV_MODELE.xltx",
        TEMPLATES_DIR / "PV_MODELE",
    ]

    for path in candidates:
        if path.exists():
            return path

    for path in TEMPLATES_DIR.iterdir():
        if path.is_file() and path.stem.upper() == "PV_MODELE":
            return path

    raise FileNotFoundError(
        "Aucun modèle Excel trouvé dans templates/. Nom attendu : PV_MODELE.xlsx."
    )


# =========================================================
# HELPERS - SQLITE
# =========================================================

def get_db_connection():
    conn = sqlite3.connect(SQLITE_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def parse_json_field(value, default):
    if not value:
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def dump_json_field(value):
    return json.dumps(value if value is not None else {}, ensure_ascii=False)


def init_app_db():
    """
    Initialise la base SQLite phase 1.
    Le fichier echaff.db est créé automatiquement à la racine du projet.
    """
    SQLITE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.executescript("""
            CREATE TABLE IF NOT EXISTS societes (
                id TEXT PRIMARY KEY,
                nom TEXT NOT NULL DEFAULT '',
                siret TEXT DEFAULT '',
                adresse TEXT DEFAULT '',
                code_postal TEXT DEFAULT '',
                ville TEXT DEFAULT '',
                pays TEXT DEFAULT 'France',
                telephone TEXT DEFAULT '',
                email TEXT DEFAULT '',
                representant_nom TEXT DEFAULT '',
                representant_prenom TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS profils (
                id TEXT PRIMARY KEY,
                societe_id TEXT,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                email TEXT NOT NULL,
                telephone TEXT DEFAULT '',
                role TEXT NOT NULL,
                actif INTEGER NOT NULL DEFAULT 1,
                signature_electronique TEXT DEFAULT '',
                certification TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (societe_id) REFERENCES societes(id)
            );

            CREATE TABLE IF NOT EXISTS chantiers (
                id TEXT PRIMARY KEY,
                societe_id TEXT,
                nom TEXT NOT NULL,
                reference_interne TEXT UNIQUE NOT NULL,
                adresse_complete TEXT DEFAULT '',
                batiment_zone_etage_secteur TEXT DEFAULT '',
                client_maitre_ouvrage TEXT DEFAULT '',
                date_debut TEXT DEFAULT '',
                date_fin_estimee TEXT DEFAULT '',
                date_fin_reelle TEXT DEFAULT '',
                statut TEXT NOT NULL DEFAULT 'brouillon',
                societe_echafaudage_responsable TEXT DEFAULT '',
                societes_utilisatrices_autorisees TEXT NOT NULL DEFAULT '[]',
                documents_associes TEXT NOT NULL DEFAULT '[]',
                historique TEXT NOT NULL DEFAULT '[]',
                qr_token TEXT UNIQUE,
                qr_code_url TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (societe_id) REFERENCES societes(id)
            );

            CREATE TABLE IF NOT EXISTS pv_reception (
                id TEXT PRIMARY KEY,
                dossier_id TEXT UNIQUE NOT NULL,
                numero_pv TEXT NOT NULL,
                chantier_id TEXT,
                chantier_nom TEXT DEFAULT '',
                statut_document TEXT NOT NULL DEFAULT 'pv_reception',
                excel_file TEXT DEFAULT '',
                pdf_file TEXT DEFAULT '',
                json_file TEXT DEFAULT '',
                client_signature_url TEXT DEFAULT '',
                data TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (chantier_id) REFERENCES chantiers(id)
            );

            CREATE TABLE IF NOT EXISTS historique_actions (
                id TEXT PRIMARY KEY,
                societe_id TEXT,
                chantier_id TEXT,
                pv_id TEXT,
                type_action TEXT NOT NULL,
                description TEXT DEFAULT '',
                auteur TEXT DEFAULT 'system',
                metadata TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_profils_role ON profils(role);
            CREATE INDEX IF NOT EXISTS idx_chantiers_reference ON chantiers(reference_interne);
            CREATE INDEX IF NOT EXISTS idx_chantiers_statut ON chantiers(statut);
            CREATE INDEX IF NOT EXISTS idx_pv_chantier_id ON pv_reception(chantier_id);
        """)
        conn.commit()

    get_or_create_main_societe_id()
    print(f"[SQLITE] Base initialisée : {SQLITE_DB_PATH}")


def get_or_create_main_societe_id() -> str:
    now = datetime.now().isoformat()
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM societes ORDER BY created_at ASC LIMIT 1")
        row = cur.fetchone()
        if row:
            return row["id"]

        societe_id = uuid4().hex
        cur.execute("""
            INSERT INTO societes (id, nom, created_at, updated_at)
            VALUES (?, '', ?, ?)
        """, (societe_id, now, now))
        conn.commit()
        return societe_id


# =========================================================
# HELPERS - FILES / JSON
# =========================================================

def save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def build_paths_for_dossier(dossier_id: str) -> dict:
    dossier_dir = DATA_DIR / dossier_id
    dossier_dir.mkdir(parents=True, exist_ok=True)

    temp_dir = dossier_dir / "tmp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    return {
        "dossier_dir": dossier_dir,
        "json_path": dossier_dir / "state.json",
        "xlsx_path": OUTPUT_DIR / f"{dossier_id}.xlsx",
        "pdf_path": OUTPUT_DIR / f"{dossier_id}.pdf",
        "temp_dir": temp_dir,
    }


def save_upload_file(upload_file: UploadFile, destination_dir: Path) -> str:
    extension = Path(upload_file.filename).suffix.lower()
    unique_filename = f"{uuid4().hex}{extension}"
    destination = destination_dir / unique_filename

    with destination.open("wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)

    relative_path = destination.relative_to(BASE_DIR)
    return str(relative_path).replace("\\", "/")


# =========================================================
# HELPERS - SIGNATURES / IMAGES
# =========================================================

def decode_base64_image(signature_data_url: str) -> bytes | None:
    if not signature_data_url or not signature_data_url.startswith("data:image"):
        return None

    try:
        _, encoded = signature_data_url.split(",", 1)
        return base64.b64decode(encoded)
    except Exception:
        return None


def save_signature_from_base64(signature_data_url: str) -> Path | None:
    image_data = decode_base64_image(signature_data_url)
    if not image_data:
        return None

    try:
        filename = f"signature_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.png"
        filepath = SIGNATURES_DIR / filename

        with filepath.open("wb") as f:
            f.write(image_data)

        return filepath
    except Exception as e:
        print("Erreur sauvegarde signature :", e)
        return None


def save_base64_signature_to_temp_png(signature_b64: str, output_path: Path) -> Path:
    image_data = decode_base64_image(signature_b64)
    if not image_data:
        raise ValueError("Signature vide ou invalide")

    image = PILImage.open(io.BytesIO(image_data)).convert("RGBA")
    background = PILImage.new("RGBA", image.size, (255, 255, 255, 255))
    image = PILImage.alpha_composite(background, image).convert("RGB")
    image.save(output_path, format="PNG")

    return output_path


def excel_col_width_to_pixels(width):
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def excel_row_height_to_pixels(height):
    if height is None:
        height = 15
    return int(height * 96 / 72)


def get_cell_or_range_bounds(ws, cell_or_range: str):
    """
    Accepte :
    - une cellule simple : AC38
    - une plage : AC38:AR38
    - une cellule appartenant à une zone fusionnée
    """
    if ":" in cell_or_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_or_range)
        return min_col, min_row, max_col, max_row

    for merged_range in ws.merged_cells.ranges:
        if cell_or_range in merged_range:
            return (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )

    col_letters = "".join(filter(str.isalpha, cell_or_range))
    row_digits = "".join(filter(str.isdigit, cell_or_range))

    col = column_index_from_string(col_letters)
    row = int(row_digits)

    return col, row, col, row


def get_range_size_pixels(ws, min_col, min_row, max_col, max_row):
    total_width = 0
    total_height = 0

    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        total_width += excel_col_width_to_pixels(width)

    for row_idx in range(min_row, max_row + 1):
        height = ws.row_dimensions[row_idx].height
        total_height += excel_row_height_to_pixels(height)

    return total_width, total_height


def insert_signature_fit_area(ws, cell_or_range: str, image_path: Path, padding_px: int = 4):
    min_col, min_row, max_col, max_row = get_cell_or_range_bounds(ws, cell_or_range)
    box_width, box_height = get_range_size_pixels(ws, min_col, min_row, max_col, max_row)

    max_width = max(20, box_width - (padding_px * 2))
    max_height = max(20, box_height - (padding_px * 2))

    with PILImage.open(image_path) as pil_img:
        img_width, img_height = pil_img.size

    ratio = min(max_width / img_width, max_height / img_height)
    final_width = max(1, int(img_width * ratio))
    final_height = max(1, int(img_height * ratio))

    offset_x = int((box_width - final_width) / 2)
    offset_y = int((box_height - final_height) / 2)

    xl_img = XLImage(str(image_path))
    xl_img.width = final_width
    xl_img.height = final_height

    marker = AnchorMarker(
        col=min_col - 1,
        row=min_row - 1,
        colOff=pixels_to_EMU(offset_x),
        rowOff=pixels_to_EMU(offset_y),
    )

    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(final_width),
        cy=pixels_to_EMU(final_height),
    )

    xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(xl_img)


# =========================================================
# HELPERS - EXCEL / PDF
# =========================================================

def export_excel_to_pdf(excel_path: Path, pdf_path: Path):
    """
    Conversion XLSX -> PDF via LibreOffice headless.
    Nécessite LibreOffice installé sur Railway.
    """
    output_dir = pdf_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        "soffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(output_dir),
        str(excel_path),
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        raise RuntimeError(
            f"Conversion PDF échouée : {result.stderr or result.stdout}"
        )

    generated_pdf = output_dir / f"{excel_path.stem}.pdf"

    if not generated_pdf.exists():
        raise RuntimeError("Le PDF n'a pas été généré par LibreOffice")

    if generated_pdf.resolve() != pdf_path.resolve():
        generated_pdf.replace(pdf_path)


def write_merged_cell(ws, cell_ref: str, value, font_size: int | None = None, bold: bool | None = None):
    target_cell = ws[cell_ref]

    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            target_cell = ws[merged_range.start_cell.coordinate]
            break

    target_cell.value = value if value is not None else ""

    if font_size is not None or bold is not None:
        new_font = copy(target_cell.font)

        if font_size is not None:
            new_font.sz = font_size

        if bold is not None:
            new_font.b = bold

        target_cell.font = new_font

    return target_cell


def clear_societes_table(ws):
    for row in range(ROW_START_SOCIETES, ROW_END_SOCIETES + 1):
        for col in [COL_SOCIETE, COL_REPRESENTANT, COL_DATE, COL_SIGNATURE]:
            cell = ws[f"{col}{row}"]
            cell.value = None

            if col == COL_SIGNATURE:
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(name="Arial", size=10, bold=False, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def validate_societes_limit(societes_utilisatrices: list):
    if len(societes_utilisatrices) > MAX_SOCIETES_UTILISATRICES:
        raise ValueError("Maximum 10 sociétés utilisatrices sur ce PV")


def fill_societes_utilisatrices_table(ws, societes_utilisatrices: list, temp_dir: Path):
    validate_societes_limit(societes_utilisatrices)
    clear_societes_table(ws)
    temp_dir.mkdir(parents=True, exist_ok=True)

    for index, soc in enumerate(societes_utilisatrices):
        row = ROW_START_SOCIETES + index

        cell_societe = f"{COL_SOCIETE}{row}"
        cell_representant = f"{COL_REPRESENTANT}{row}"
        cell_date = f"{COL_DATE}{row}"
        cell_signature = f"{COL_SIGNATURE}{row}"

        ws[cell_societe] = soc.get("societe", "")
        ws[cell_representant] = soc.get("representant", "")

        signed = bool(soc.get("signed", False))
        signature_b64 = soc.get("signature_b64", "")

        if signed and signature_b64:
            date_signature = soc.get("date_signature", "")
            heure_signature = soc.get("heure_signature", "")
            ws[cell_date] = f"{date_signature} {heure_signature}".strip()

            temp_signature_path = temp_dir / f"signature_societe_{row}.png"
            save_base64_signature_to_temp_png(signature_b64, temp_signature_path)
            insert_signature_fit_area(ws, cell_signature, temp_signature_path)
        else:
            ws[cell_date] = ""
            ws[cell_signature].fill = ATTENTE_FILL
            ws[cell_signature].font = ATTENTE_FONT
            ws[cell_signature].alignment = ATTENTE_ALIGNMENT
            ws[cell_signature] = "EN ATTENTE"


def apply_page_setup(ws):
    ws.print_area = PRINT_AREA


# =========================================================
# EMAIL
# =========================================================

def send_email(destinataires: str, sujet: str, contenu: str):
    if not destinataires:
        return

    if not SMTP_USER or not SMTP_PASSWORD or not SMTP_FROM:
        print("Email non envoyé : variables SMTP manquantes")
        return

    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_FROM
        msg["To"] = destinataires
        msg["Subject"] = sujet
        msg.attach(MIMEText(contenu, "plain", "utf-8"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)

    except Exception as e:
        print("Erreur envoi mail :", e)


# =========================================================
# MAPPING EXCEL
# =========================================================

TEXT_CELL_MAP = {
    "chantier": "B4",
    "adresse": "B5",
    "date_montage": "B6",
    "maitre_ouvrage": "B8",
    "contact_mo": "B9",
    "tel_mo": "T9",
    "entreprise_montage": "B11",
    "contact_montage": "B12",
    "tel_montage": "T12",
    "entreprise_utilisatrice": "B13",
    "contact_utilisatrice": "B14",
    "tel_utilisatrice": "T14",
    "echafaudages_speciaux": "B19",
    "restriction_utilisation": "B24",
}

TYPE_ECHAFAUDAGE_MAP = {
    "type_facade": "A16",
    "type_recueil": "A17",
    "type_filet": "A18",
    "type_bache": "H16",
    "type_plateforme": "H17",
    "type_escaliers": "H18",
    "type_toit": "O16",
    "type_toiture": "O17",
}

CLASSE_CHARGE_MAP = {
    "150": "D21",
    "200": "I21",
    "300": "N21",
    "450": "T21",
    "600": "T22",
}

CLASSE_LARGEUR_MAP = {
    "W06": "H23",
    "W09": "N23",
    "W": "T23",
}

CHECKLIST_MAP = {
    "q_apparentement_intacts": {"oui": "AP4", "non": "AQ4", "na": "AR4"},
    "q_resistance_support": {"oui": "AP6", "non": "AQ6", "na": "AR6"},
    "q_verins_reglage": {"oui": "AP7", "non": "AQ7", "na": "AR7"},
    "q_contreventements": {"oui": "AP8", "non": "AQ8", "na": "AR8"},
    "q_traverses_longitudinales": {"oui": "AP9", "non": "AQ9", "na": "AR9"},
    "q_poutres_treillis": {"oui": "AP10", "non": "AQ10", "na": "AR10"},
    "q_ancrages_nombre": {"value": "AP11"},
    "q_niveaux_recouverts": {"oui": "AP12", "non": "AQ12", "na": "AR12"},
    "q_planchers_compris": {"oui": "AP13", "non": "AQ13", "na": "AR13"},
    "q_au_niveau_des_angles": {"oui": "AP14", "non": "AQ14", "na": "AR14"},
    "q_madriers": {"oui": "AP15", "non": "AQ15", "na": "AR15"},
    "q_ouvertures": {"oui": "AP16", "non": "AQ16", "na": "AR16"},
    "q_dispositifs_securite": {"oui": "AP17", "non": "AQ17", "na": "AR17"},
    "q_distance_mur": {"oui": "AP18", "non": "AQ18", "na": "AR18"},
    "q_garde_corps_interieur": {"oui": "AP19", "non": "AQ19", "na": "AR19"},
    "q_montees_acces": {"oui": "AP20", "non": "AQ20", "na": "AR20"},
    "q_tour_escaliers": {"oui": "AP21", "non": "AQ21", "na": "AR21"},
    "q_echelle_appui": {"oui": "AP22", "non": "AQ22", "na": "AR22"},
    "q_exigences_recueil": {"oui": "AP23", "non": "AQ23", "na": "AR23"},
    "q_conduites_tension": {"oui": "AP24", "non": "AQ24", "na": "AR24"},
    "q_ecran_protection": {"oui": "AP25", "non": "AQ25", "na": "AR25"},
    "q_toit_protection_ctrl": {"oui": "AP26", "non": "AQ26", "na": "AR26"},
    "q_securite_circulation": {"oui": "AP27", "non": "AQ27", "na": "AR27"},
    "q_aux_acces": {"oui": "AP28", "non": "AQ28", "na": "AR28"},
    "q_clotures": {"oui": "AP29", "non": "AQ29", "na": "AR29"},
}


def mark_x(ws, cell_ref: str):
    cell = write_merged_cell(ws, cell_ref, "X", font_size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def fill_simple_text_fields(ws, dossier_data: dict):
    for payload_key, cell_ref in TEXT_CELL_MAP.items():
        value = dossier_data.get(payload_key, "")
        write_merged_cell(ws, cell_ref, value)


def fill_type_echafaudage_fields(ws, dossier_data: dict):
    for payload_key, cell_ref in TYPE_ECHAFAUDAGE_MAP.items():
        if dossier_data.get(payload_key, False):
            mark_x(ws, cell_ref)


def fill_type_entreprise_field(ws, dossier_data: dict):
    value = str(dossier_data.get("type_entreprise", "")).strip().lower()

    if value == "montage":
        texte = "Entreprise de montage"
    elif value == "propre":
        texte = "Entreprise de montage pour usage propre"
    else:
        texte = ""

    if texte:
        cell = write_merged_cell(ws, TYPE_ENTREPRISE_TEXT_CELL, texte)
        cell.font = Font(name="Calibri", size=18, bold=False)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fill_classe_charge(ws, dossier_data: dict):
    value = str(dossier_data.get("classe_charge", "")).strip()
    cell_ref = CLASSE_CHARGE_MAP.get(value)
    if cell_ref:
        mark_x(ws, cell_ref)


def fill_classe_largeur(ws, dossier_data: dict):
    value = str(dossier_data.get("classe_largeur", "")).strip().upper()
    cell_ref = CLASSE_LARGEUR_MAP.get(value)
    if cell_ref:
        mark_x(ws, cell_ref)

    if value == "W":
        largeur_libre = str(dossier_data.get("largeur_libre", "")).strip()
        if largeur_libre:
            current = ws["T23"].value or ""
            ws["T23"] = f"X ({largeur_libre})" if current == "X" else largeur_libre


def fill_checklist_fields(ws, dossier_data: dict):
    for payload_key, mapping in CHECKLIST_MAP.items():
        value = dossier_data.get(payload_key, "")

        if "value" in mapping:
            if value not in ("", None):
                write_merged_cell(ws, mapping["value"], value)
            continue

        value = str(value).strip().lower()
        cell_ref = mapping.get(value)

        if cell_ref:
            mark_x(ws, cell_ref)


def fill_observations_block(ws, dossier_data: dict):
    observations = dossier_data.get("observations", "")
    cell_obs = write_merged_cell(ws, OBSERVATIONS_CELL, observations, font_size=18)
    cell_obs.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def fill_verificateur_block(ws, dossier_data: dict):
    verificateur_nom = dossier_data.get("verificateur_nom", "").strip()
    verificateur_numero_diplome = dossier_data.get("verificateur_numero_diplome", "").strip()
    verificateur_lien_diplome = dossier_data.get("verificateur_lien_diplome", "").strip()

    if verificateur_nom:
        write_merged_cell(ws, VERIF_NAME_CELL, verificateur_nom)

    if verificateur_numero_diplome:
        cell_diplome = write_merged_cell(ws, VERIF_DIPLOME_CELL, verificateur_numero_diplome)

        if verificateur_lien_diplome:
            full_path = (BASE_DIR / verificateur_lien_diplome.strip("/")).resolve()
            if full_path.exists():
                cell_diplome.hyperlink = str(full_path)
                cell_diplome.value = verificateur_numero_diplome
                cell_diplome.style = "Hyperlink"

    verification_datetime = dossier_data.get("verification_datetime")
    if verification_datetime:
        try:
            dt = datetime.fromisoformat(verification_datetime)
        except ValueError:
            dt = datetime.now()
    else:
        dt = datetime.now()
        dossier_data["verification_datetime"] = dt.isoformat()

    write_merged_cell(ws, VERIF_DATE_CELL, dt.strftime("%d/%m/%Y"))
    write_merged_cell(ws, VERIF_HOUR_CELL, dt.strftime("%H:%M"))

    signature_data = dossier_data.get("signature", "")
    signature_path = save_signature_from_base64(signature_data)

    if signature_path and signature_path.exists():
        try:
            insert_signature_fit_area(ws, VERIF_SIGNATURE_CELL, signature_path, padding_px=4)
        except Exception as e:
            print("Erreur insertion signature vérificateur Excel :", e)


def fill_client_block(ws, dossier_data: dict):
    client = dossier_data.get("client_signature", {}) or {}

    client_nom = client.get("nom_signataire", "").strip()
    client_signature = client.get("signature_b64", "")
    client_datetime = client.get("signature_datetime")

    if client_nom:
        write_merged_cell(ws, CLIENT_NAME_CELL, client_nom)

    if client_datetime:
        try:
            dt = datetime.fromisoformat(client_datetime)
        except ValueError:
            dt = None
    else:
        dt = None

    if dt:
        write_merged_cell(ws, CLIENT_DATE_CELL, dt.strftime("%d/%m/%Y"))
        write_merged_cell(ws, CLIENT_HOUR_CELL, dt.strftime("%H:%M"))

    if client_signature:
        signature_path = save_signature_from_base64(client_signature)
        if signature_path and signature_path.exists():
            try:
                insert_signature_fit_area(ws, CLIENT_SIGNATURE_CELL, signature_path, padding_px=4)
            except Exception as e:
                print("Erreur insertion signature client Excel :", e)


def regenerate_excel_from_data(output_xlsx_path: Path, dossier_data: dict, temp_dir: Path):
    template_path = find_excel_template()
    shutil.copy2(template_path, output_xlsx_path)

    wb = load_workbook(output_xlsx_path)
    ws = wb["Formulaire"] if "Formulaire" in wb.sheetnames else wb.active

    numero_pv = dossier_data.get("numero_pv") or datetime.now().strftime("%Y%m%d%H%M%S")
    dossier_data["numero_pv"] = numero_pv

    ws["A1"] = f"PROCÈS-VERBAL DE CONTRÔLE N°{numero_pv}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    fill_simple_text_fields(ws, dossier_data)
    fill_type_echafaudage_fields(ws, dossier_data)
    fill_classe_charge(ws, dossier_data)
    fill_classe_largeur(ws, dossier_data)
    fill_checklist_fields(ws, dossier_data)
    fill_type_entreprise_field(ws, dossier_data)
    fill_observations_block(ws, dossier_data)
    fill_verificateur_block(ws, dossier_data)
    fill_client_block(ws, dossier_data)
    fill_societes_utilisatrices_table(ws, dossier_data.get("societes_utilisatrices", []), temp_dir)
    apply_page_setup(ws)

    wb.save(output_xlsx_path)


# =========================================================
# HELPERS - BUSINESS
# =========================================================

def get_diplome_status(date_echeance_str: str | None):
    if not date_echeance_str:
        return {"label": "Date manquante", "color": "grey"}

    try:
        echeance = datetime.strptime(date_echeance_str, "%Y-%m-%d").date()
    except ValueError:
        return {"label": "Date invalide", "color": "grey"}

    today = date.today()
    delta_days = (echeance - today).days

    if delta_days < 0:
        return {"label": "Expiré", "color": "red"}
    if delta_days <= 183:
        return {"label": "Renouvellement < 6 mois", "color": "orange"}
    return {"label": "Valide", "color": "green"}


def prepare_pv_payload(data: dict) -> dict:
    data = enrich_pv_data_from_chantier(dict(data))
    dossier_id = data.get("dossier_id") or f"pv_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    numero_pv = data.get("numero_pv") or datetime.now().strftime("%Y%m%d%H%M%S")

    payload = {
        "dossier_id": dossier_id,
        "numero_pv": numero_pv,
        "chantier_id": data.get("chantier_id", "").strip(),
        "chantier": data.get("chantier", ""),
        "adresse": data.get("adresse", ""),
        "date_montage": data.get("date_montage", ""),
        "observations": data.get("observations", ""),
        "maitre_ouvrage": data.get("maitre_ouvrage", ""),
        "contact_mo": data.get("contact_mo", ""),
        "tel_mo": data.get("tel_mo", ""),
        "entreprise_montage": data.get("entreprise_montage", ""),
        "contact_montage": data.get("contact_montage", ""),
        "tel_montage": data.get("tel_montage", ""),
        "entreprise_utilisatrice": data.get("entreprise_utilisatrice", ""),
        "contact_utilisatrice": data.get("contact_utilisatrice", ""),
        "tel_utilisatrice": data.get("tel_utilisatrice", ""),
        "type_facade": data.get("type_facade", False),
        "type_recueil": data.get("type_recueil", False),
        "type_filet": data.get("type_filet", False),
        "type_bache": data.get("type_bache", False),
        "type_plateforme": data.get("type_plateforme", False),
        "type_escaliers": data.get("type_escaliers", False),
        "type_toit": data.get("type_toit", False),
        "type_toiture": data.get("type_toiture", False),
        "type_entreprise": data.get("type_entreprise", ""),
        "echafaudages_speciaux": data.get("echafaudages_speciaux", ""),
        "classe_charge": data.get("classe_charge", ""),
        "classe_largeur": data.get("classe_largeur", ""),
        "largeur_libre": data.get("largeur_libre", ""),
        "restriction_utilisation": data.get("restriction_utilisation", ""),
        "verificateur_nom": data.get("verificateur_nom", "").strip(),
        "verificateur_prenom": data.get("verificateur_prenom", "").strip(),
        "verificateur_email": data.get("verificateur_email", "").strip(),
        "verificateur_telephone": data.get("verificateur_telephone", "").strip(),
        "verificateur_numero_diplome": data.get("verificateur_numero_diplome", "").strip(),
        "verificateur_lien_diplome": data.get("verificateur_lien_diplome", "").strip(),
        "verificateur_statut_color": data.get("verificateur_statut_color", "").strip(),
        "verificateur_statut_label": data.get("verificateur_statut_label", "").strip(),
        "verificateur_date_echeance": data.get("verificateur_date_echeance", "").strip(),
        "signature": data.get("signature", ""),
        "verification_datetime": data.get("verification_datetime") or datetime.now().isoformat(),
        "client_signature": data.get("client_signature", {}),
        "societes_utilisatrices": normalize_societes_utilisatrices_from_payload(data),
        "email_utilisatrice": data.get("email_utilisatrice", "").strip(),
        "email_mo": data.get("email_mo", "").strip(),
    }

    for key in CHECKLIST_MAP.keys():
        payload[key] = data.get(key, "")

    validate_societes_limit(payload["societes_utilisatrices"])
    return payload


def apply_client_signature_payload(dossier_data: dict, payload: dict) -> dict:
    societes_utilisatrices = payload.get("societes_utilisatrices", [])
    validate_societes_limit(societes_utilisatrices)

    dossier_data["client_signature"] = {
        "nom_signataire": payload.get("client_nom_signataire", "").strip(),
        "email": payload.get("client_email", "").strip(),
        "telephone": payload.get("client_telephone", "").strip(),
        "signature_b64": payload.get("client_signature", ""),
        "signature_datetime": datetime.now().isoformat(),
    }

    dossier_data["societes_utilisatrices"] = societes_utilisatrices
    return dossier_data


def regenerate_pv_files(dossier_data: dict) -> dict:
    dossier_id = dossier_data["dossier_id"]
    paths = build_paths_for_dossier(dossier_id)

    # On conserve le JSON de travail pour compatibilité avec l'existant.
    # Les données métier sont aussi sauvegardées dans PostgreSQL via save_pv_reception_to_db.
    save_json(paths["json_path"], dossier_data)

    regenerate_excel_from_data(
        output_xlsx_path=paths["xlsx_path"],
        dossier_data=dossier_data,
        temp_dir=paths["temp_dir"],
    )

    pdf_generated = True
    try:
        export_excel_to_pdf(paths["xlsx_path"], paths["pdf_path"])
    except Exception as e:
        traceback.print_exc()
        print("Erreur export PDF :", e)
        pdf_generated = False

    generated = {
        "json_path": paths["json_path"],
        "xlsx_path": paths["xlsx_path"],
        "pdf_path": paths["pdf_path"] if pdf_generated else None,
    }

    save_pv_reception_to_db(dossier_data, generated)
    return generated


# =========================================================
# EXTENSION ECHAFF - PHASE 1
# Ajout sans modification de la structure existante
# Objectif : informations société, profils société, chantiers société
# =========================================================

# Compatibilité interne : ces constantes restent en place pour ne pas casser la structure.
# Elles ne pointent plus vers des fichiers : elles servent de clés logiques pour les helpers PostgreSQL.
ECHAFF_SOCIETE_FILE = Path("postgres://societe")
ECHAFF_PROFILS_FILE = Path("postgres://profils")
ECHAFF_CHANTIERS_FILE = Path("postgres://chantiers")


ROLES_ECHAFF = [
    "admin_societe_echafaudage",
    "responsable_certifie",
    "collaborateur_echafaudage",
    "societe_utilisatrice",
    "inspecteur_lecture_seule",
    "visiteur_qr_code",
]


STATUTS_CHANTIER = [
    "brouillon",
    "en_cours",
    "en_attente_signatures",
    "actif",
    "cloture",
    "archive",
]


def load_list_json(path: Path) -> list:
    key = str(path)

    with get_db_connection() as conn:
        cur = conn.cursor()

        if key == str(ECHAFF_PROFILS_FILE):
            cur.execute("""
                SELECT p.*, s.nom AS societe
                FROM profils p
                LEFT JOIN societes s ON s.id = p.societe_id
                ORDER BY p.created_at DESC
            """)
            rows = cur.fetchall()
            profils = []
            for row in rows:
                item = dict(row)
                item["actif"] = bool(item.get("actif"))
                item["certification"] = parse_json_field(item.get("certification"), {})
                profils.append(item)
            return profils

        if key == str(ECHAFF_CHANTIERS_FILE):
            cur.execute("""
                SELECT c.*, s.nom AS societe_nom
                FROM chantiers c
                LEFT JOIN societes s ON s.id = c.societe_id
                ORDER BY c.created_at DESC
            """)
            rows = cur.fetchall()
            chantiers = []
            for row in rows:
                item = dict(row)
                item["societes_utilisatrices_autorisees"] = parse_json_field(item.get("societes_utilisatrices_autorisees"), [])
                item["documents_associes"] = parse_json_field(item.get("documents_associes"), [])
                item["historique"] = parse_json_field(item.get("historique"), [])
                chantiers.append(item)
            return chantiers

    return []


def save_list_json(path: Path, data: list) -> None:
    key = str(path)
    societe_id = get_or_create_main_societe_id()
    now = datetime.now().isoformat()

    with get_db_connection() as conn:
        cur = conn.cursor()

        if key == str(ECHAFF_PROFILS_FILE):
            cur.execute("DELETE FROM profils")
            for item in data:
                cur.execute("""
                    INSERT INTO profils (
                        id, societe_id, nom, prenom, email, telephone, role, actif,
                        signature_electronique, certification, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item.get("id") or uuid4().hex,
                    societe_id,
                    item.get("nom", ""),
                    item.get("prenom", ""),
                    item.get("email", ""),
                    item.get("telephone", ""),
                    item.get("role", ""),
                    1 if item.get("actif", True) else 0,
                    item.get("signature_electronique", ""),
                    dump_json_field(item.get("certification", {}) or {}),
                    item.get("created_at") or now,
                    now,
                ))

        elif key == str(ECHAFF_CHANTIERS_FILE):
            cur.execute("DELETE FROM chantiers")
            for item in data:
                cur.execute("""
                    INSERT INTO chantiers (
                        id, societe_id, nom, reference_interne, adresse_complete,
                        batiment_zone_etage_secteur, client_maitre_ouvrage,
                        date_debut, date_fin_estimee, date_fin_reelle, statut,
                        societe_echafaudage_responsable, societes_utilisatrices_autorisees,
                        documents_associes, historique, qr_token, qr_code_url,
                        created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item.get("id") or uuid4().hex,
                    societe_id,
                    item.get("nom", ""),
                    item.get("reference_interne") or generate_next_chantier_reference(),
                    item.get("adresse_complete", ""),
                    item.get("batiment_zone_etage_secteur", ""),
                    item.get("client_maitre_ouvrage", ""),
                    item.get("date_debut") or "",
                    item.get("date_fin_estimee") or "",
                    item.get("date_fin_reelle") or "",
                    item.get("statut", "brouillon"),
                    item.get("societe_echafaudage_responsable", ""),
                    dump_json_field(item.get("societes_utilisatrices_autorisees", []) or []),
                    dump_json_field(item.get("documents_associes", []) or []),
                    dump_json_field(item.get("historique", []) or []),
                    item.get("qr_token"),
                    item.get("qr_code_url", ""),
                    item.get("created_at") or now,
                    now,
                ))

        conn.commit()


def load_dict_json(path: Path) -> dict:
    key = str(path)
    if key != str(ECHAFF_SOCIETE_FILE):
        return {}

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM societes ORDER BY created_at ASC LIMIT 1")
        row = cur.fetchone()
        return dict(row) if row else {}


def save_dict_json(path: Path, data: dict) -> None:
    key = str(path)
    if key != str(ECHAFF_SOCIETE_FILE):
        return

    now = datetime.now().isoformat()
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id FROM societes ORDER BY created_at ASC LIMIT 1")
        row = cur.fetchone()

        if row:
            cur.execute("""
                UPDATE societes SET
                    nom=?, siret=?, adresse=?, code_postal=?, ville=?, pays=?,
                    telephone=?, email=?, representant_nom=?, representant_prenom=?,
                    updated_at=?
                WHERE id=?
            """, (
                data.get("nom", ""),
                data.get("siret", ""),
                data.get("adresse", ""),
                data.get("code_postal", ""),
                data.get("ville", ""),
                data.get("pays", "France"),
                data.get("telephone", ""),
                data.get("email", ""),
                data.get("representant_nom", ""),
                data.get("representant_prenom", ""),
                now,
                row["id"],
            ))
        else:
            cur.execute("""
                INSERT INTO societes (
                    id, nom, siret, adresse, code_postal, ville, pays, telephone, email,
                    representant_nom, representant_prenom, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                uuid4().hex,
                data.get("nom", ""),
                data.get("siret", ""),
                data.get("adresse", ""),
                data.get("code_postal", ""),
                data.get("ville", ""),
                data.get("pays", "France"),
                data.get("telephone", ""),
                data.get("email", ""),
                data.get("representant_nom", ""),
                data.get("representant_prenom", ""),
                now,
                now,
            ))

        conn.commit()


def append_historique_chantier(chantier: dict, action: str, auteur: str = "system") -> dict:
    historique = chantier.get("historique", [])
    historique.append({
        "date": datetime.now().isoformat(),
        "action": action,
        "auteur": auteur,
    })
    chantier["historique"] = historique
    return chantier


def get_chantier_by_id(chantier_id: str) -> dict | None:
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    for chantier in chantiers:
        if chantier.get("id") == chantier_id:
            return chantier
    return None


def save_chantier(updated_chantier: dict) -> dict:
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    for index, chantier in enumerate(chantiers):
        if chantier.get("id") == updated_chantier.get("id"):
            updated_chantier["updated_at"] = datetime.now().isoformat()
            chantiers[index] = updated_chantier
            save_list_json(ECHAFF_CHANTIERS_FILE, chantiers)
            return updated_chantier

    raise HTTPException(status_code=404, detail="Chantier introuvable")


def delete_chantier_by_id(chantier_id: str) -> bool:
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    new_chantiers = [c for c in chantiers if c.get("id") != chantier_id]

    if len(new_chantiers) == len(chantiers):
        return False

    save_list_json(ECHAFF_CHANTIERS_FILE, new_chantiers)
    return True


def get_pvs_for_chantier(chantier_id: str) -> list:
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT dossier_id, numero_pv, chantier_nom AS chantier, created_at,
                   excel_file, pdf_file, client_signature_url, statut_document AS statut
            FROM pv_reception
            WHERE chantier_id = ?
            ORDER BY created_at DESC
        """, (chantier_id,))
        return [dict(row) for row in cur.fetchall()]


def save_pv_reception_to_db(dossier_data: dict, generated: dict) -> None:
    chantier_id = dossier_data.get("chantier_id") or None
    now = datetime.now().isoformat()

    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO pv_reception (
                id, dossier_id, numero_pv, chantier_id, chantier_nom, statut_document,
                excel_file, pdf_file, json_file, client_signature_url, data, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(dossier_id) DO UPDATE SET
                numero_pv = excluded.numero_pv,
                chantier_id = excluded.chantier_id,
                chantier_nom = excluded.chantier_nom,
                statut_document = excluded.statut_document,
                excel_file = excluded.excel_file,
                pdf_file = excluded.pdf_file,
                json_file = excluded.json_file,
                client_signature_url = excluded.client_signature_url,
                data = excluded.data,
                updated_at = excluded.updated_at
        """, (
            uuid4().hex,
            dossier_data.get("dossier_id"),
            dossier_data.get("numero_pv"),
            chantier_id,
            dossier_data.get("chantier", ""),
            dossier_data.get("statut_document", "pv_reception"),
            f"/output/{generated['xlsx_path'].name}" if generated.get("xlsx_path") else "",
            f"/output/{generated['pdf_path'].name}" if generated.get("pdf_path") else "",
            f"/data/{dossier_data.get('dossier_id')}/state.json",
            f"/client-signature/{dossier_data.get('dossier_id')}",
            dump_json_field(dossier_data),
            now,
            now,
        ))
        conn.commit()


def get_notifications_chantier(chantier_id: str) -> list:
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        return []

    notifications = []
    pvs = get_pvs_for_chantier(chantier_id)

    if not pvs:
        notifications.append({
            "type": "warning",
            "label": "Aucun PV de réception généré pour ce chantier",
        })

    for pv in pvs:
        dossier_id = pv.get("dossier_id")
        state_path = DATA_DIR / dossier_id / "state.json"
        if not state_path.exists():
            continue

        data = load_json(state_path)
        societes = data.get("societes_utilisatrices", [])
        pending = [s for s in societes if not s.get("signed")]

        if pending:
            notifications.append({
                "type": "danger",
                "label": f"{len(pending)} signature(s) en attente pour le PV {pv.get('numero_pv')}",
            })

    if chantier.get("statut") == "archive":
        notifications.append({
            "type": "info",
            "label": "Chantier archivé",
        })

    return notifications


def get_global_notifications() -> list:
    notifications = []
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)

    for chantier in chantiers:
        chantier_notifications = get_notifications_chantier(chantier.get("id", ""))
        for notif in chantier_notifications:
            notifications.append({
                "chantier_id": chantier.get("id"),
                "chantier_nom": chantier.get("nom"),
                **notif,
            })

    return notifications


# ---------------------------------------------------------
# EXTENSION ECHAFF - PHASE 1 : société unique
# ---------------------------------------------------------

def get_current_societe() -> dict:
    """
    Phase 1 : une seule société d'échafaudage.
    Tous les profils et chantiers sont rattachés à cette société.
    """
    return load_dict_json(ECHAFF_SOCIETE_FILE)


def get_current_societe_name() -> str:
    societe = get_current_societe()
    return societe.get("nom", "")


def generate_next_chantier_reference() -> str:
    """
    Génère une référence unique incrémentée automatiquement.
    Format : CH-0001, CH-0002, etc.
    """
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    max_number = 0

    for chantier in chantiers:
        ref = str(chantier.get("reference_interne", ""))
        if ref.startswith("CH-"):
            try:
                number = int(ref.replace("CH-", ""))
                max_number = max(max_number, number)
            except ValueError:
                continue

    return f"CH-{max_number + 1:04d}"


def get_default_verificateur_certifie() -> dict | None:
    """
    Retourne le premier profil certifié actif de la société.
    Sert au pré-remplissage du PV si aucun vérificateur n'est choisi côté formulaire.
    """
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    societe_nom = get_current_societe_name()

    for profil in profils:
        certification = profil.get("certification", {}) or {}
        if (
            profil.get("actif", True)
            and profil.get("societe", "") == societe_nom
            and certification.get("certifie")
        ):
            return profil

    return None


def enrich_pv_data_from_chantier(data: dict) -> dict:
    """
    Si un PV est associé à un chantier, on complète automatiquement :
    - chantier
    - adresse
    - maître d'ouvrage
    - entreprise de montage
    - téléphone montage
    - vérificateur certifié par défaut si besoin
    """
    chantier_id = data.get("chantier_id", "")
    if not chantier_id:
        return data

    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        return data

    societe = get_current_societe()
    verificateur = get_default_verificateur_certifie()

    data.setdefault("chantier", chantier.get("nom", ""))
    data.setdefault("adresse", chantier.get("adresse_complete", ""))
    data.setdefault("maitre_ouvrage", chantier.get("client_maitre_ouvrage", ""))

    if not data.get("chantier"):
        data["chantier"] = chantier.get("nom", "")
    if not data.get("adresse"):
        data["adresse"] = chantier.get("adresse_complete", "")
    if not data.get("maitre_ouvrage"):
        data["maitre_ouvrage"] = chantier.get("client_maitre_ouvrage", "")

    if not data.get("entreprise_montage"):
        data["entreprise_montage"] = societe.get("nom", "")
    if not data.get("contact_montage"):
        data["contact_montage"] = f"{societe.get('representant_prenom', '')} {societe.get('representant_nom', '')}".strip()
    if not data.get("tel_montage"):
        data["tel_montage"] = societe.get("telephone", "")

    if verificateur and not data.get("verificateur_nom"):
        data["verificateur_nom"] = f"{verificateur.get('nom', '')} {verificateur.get('prenom', '')}".strip()
        data["verificateur_prenom"] = verificateur.get("prenom", "")
        data["verificateur_email"] = verificateur.get("email", "")
        data["verificateur_telephone"] = verificateur.get("telephone", "")
        certification = verificateur.get("certification", {}) or {}
        data["verificateur_numero_diplome"] = certification.get("reference", "")
        data["verificateur_date_echeance"] = certification.get("date_validite", "")

    return data


def normalize_societes_utilisatrices_from_payload(data: dict) -> list:
    """
    Accepte :
    - une liste societes_utilisatrices envoyée par le nouveau formulaire
    - une liste de profils sélectionnés via societes_utilisatrices_profils_ids
    - ou l'ancien format entreprise_utilisatrice/contact/tel/email
    """
    selected_profile_ids = data.get("societes_utilisatrices_profils_ids", [])
    if isinstance(selected_profile_ids, list) and selected_profile_ids:
        profils = load_list_json(ECHAFF_PROFILS_FILE)
        selected = []

        for profil in profils:
            if profil.get("id") in selected_profile_ids and profil.get("role") == "societe_utilisatrice":
                selected.append({
                    "profil_id": profil.get("id"),
                    "societe": profil.get("societe", ""),
                    "representant": f"{profil.get('prenom', '')} {profil.get('nom', '')}".strip(),
                    "telephone": profil.get("telephone", ""),
                    "email": profil.get("email", ""),
                    "signed": False,
                    "date_signature": "",
                    "heure_signature": "",
                    "signature_b64": "",
                })

        return selected[:MAX_SOCIETES_UTILISATRICES]

    societes = data.get("societes_utilisatrices", [])
    if isinstance(societes, list) and societes:
        return societes[:MAX_SOCIETES_UTILISATRICES]

    societe = data.get("entreprise_utilisatrice", "")
    contact = data.get("contact_utilisatrice", "")
    telephone = data.get("tel_utilisatrice", "")
    email = data.get("email_utilisatrice", "")

    if not any([societe, contact, telephone, email]):
        return []

    return [{
        "societe": societe,
        "representant": contact,
        "telephone": telephone,
        "email": email,
        "signed": False,
        "date_signature": "",
        "heure_signature": "",
        "signature_b64": "",
    }]


def generate_qr_code_for_chantier(chantier_id: str) -> str:
    if qrcode is None:
        raise RuntimeError("Le package qrcode n'est pas installé. Ajoutez qrcode[pil] dans requirements.txt")

    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    qr_url = f"{APP_PUBLIC_URL}/qr/chantier/{chantier_id}"
    filename = f"chantier_{chantier_id}.png"
    output_path = QR_CODES_DIR / filename

    img = qrcode.make(qr_url)
    img.save(output_path)

    return f"/uploads/qr_codes/{filename}"


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_echaff(request: Request):
    """
    Page d'accueil phase 1 avec menu :
    - informations société : création / mise à jour
    - liste des chantiers : création, modification, changement de statut, suppression, archivage
    - accès aux PV de réception par chantier
    """
    societe = load_dict_json(ECHAFF_SOCIETE_FILE)
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)

    chantiers_actifs = [c for c in chantiers if c.get("statut") != "archive"]
    chantiers_archives = [c for c in chantiers if c.get("statut") == "archive"]

    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context={
            "request": request,
            "societe": societe,
            "nb_profils": len(profils),
            "nb_chantiers": len(chantiers),
            "nb_chantiers_actifs": len(chantiers_actifs),
            "nb_chantiers_archives": len(chantiers_archives),
            "chantiers": chantiers_actifs,
            "chantiers_archives": chantiers_archives,
            "notifications": get_global_notifications(),
            "roles": ROLES_ECHAFF,
            "statuts_chantier": STATUTS_CHANTIER,
            "menu": [
                {
                    "label": "Informations société",
                    "description": "Créer ou mettre à jour les informations de la société d’échafaudage.",
                    "url": "/societe",
                },
                {
                    "label": "Liste des chantiers",
                    "description": "Créer un chantier, modifier son statut, le supprimer ou l’archiver.",
                    "url": "/chantiers",
                },
                {
                    "label": "Profils société",
                    "description": "Gérer les utilisateurs, rôles et certifications.",
                    "url": "/profils",
                },
            ],
        }
    )


@app.get("/accueil", response_class=HTMLResponse)
def accueil_echaff(request: Request):
    return dashboard_echaff(request)


# ---------------------------------------------------------
# EXTENSION ECHAFF - INFORMATIONS SOCIETE
# ---------------------------------------------------------

@app.get("/societe", response_class=HTMLResponse)
def societe_form(request: Request):
    societe = load_dict_json(ECHAFF_SOCIETE_FILE)
    return templates.TemplateResponse(
        request=request,
        name="societe_form.html",
        context={"request": request, "societe": societe}
    )


@app.post("/societe", response_class=HTMLResponse)
async def societe_save(
    request: Request,
    nom: str = Form(...),
    siret: str = Form(""),
    adresse: str = Form(""),
    code_postal: str = Form(""),
    ville: str = Form(""),
    pays: str = Form("France"),
    telephone: str = Form(""),
    email: str = Form(""),
    representant_nom: str = Form(""),
    representant_prenom: str = Form(""),
):
    societe = {
        "id": "societe_principale_phase1",
        "nom": nom.strip(),
        "siret": siret.strip(),
        "adresse": adresse.strip(),
        "code_postal": code_postal.strip(),
        "ville": ville.strip(),
        "pays": pays.strip(),
        "telephone": telephone.strip(),
        "email": email.strip(),
        "representant_nom": representant_nom.strip(),
        "representant_prenom": representant_prenom.strip(),
        "updated_at": datetime.now().isoformat(),
    }

    save_dict_json(ECHAFF_SOCIETE_FILE, societe)
    return RedirectResponse("/societe", status_code=303)


@app.get("/api/societe")
def api_get_societe():
    return load_dict_json(ECHAFF_SOCIETE_FILE)


@app.post("/api/societe")
async def api_save_societe(request: Request):
    payload = await request.json()
    payload["id"] = payload.get("id") or "societe_principale_phase1"
    payload["updated_at"] = datetime.now().isoformat()
    save_dict_json(ECHAFF_SOCIETE_FILE, payload)
    return {"success": True, "societe": payload}


# ---------------------------------------------------------
# EXTENSION ECHAFF - PROFILS SOCIETE
# ---------------------------------------------------------

@app.get("/profils", response_class=HTMLResponse)
def profils_liste(request: Request):
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    societe = get_current_societe()
    return templates.TemplateResponse(
        request=request,
        name="profils_liste.html",
        context={
            "request": request,
            "profils": profils,
            "roles": ROLES_ECHAFF,
            "societe": societe,
        }
    )


def get_profil_by_id(profil_id: str) -> dict | None:
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    for profil in profils:
        if profil.get("id") == profil_id:
            return profil
    return None


def save_profil(updated_profil: dict) -> dict:
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    for index, profil in enumerate(profils):
        if profil.get("id") == updated_profil.get("id"):
            updated_profil["updated_at"] = datetime.now().isoformat()
            profils[index] = updated_profil
            save_list_json(ECHAFF_PROFILS_FILE, profils)
            return updated_profil

    raise HTTPException(status_code=404, detail="Profil introuvable")


def delete_profil_by_id(profil_id: str) -> bool:
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    new_profils = [p for p in profils if p.get("id") != profil_id]

    if len(new_profils) == len(profils):
        return False

    save_list_json(ECHAFF_PROFILS_FILE, new_profils)
    return True


def get_profils_societes_utilisatrices() -> list:
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    return [
        profil for profil in profils
        if profil.get("role") == "societe_utilisatrice" and profil.get("actif", True)
    ]


@app.get("/profils/nouveau", response_class=HTMLResponse)
def profil_form(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="profil_form.html",
        context={
            "request": request,
            "roles": ROLES_ECHAFF,
            "societe": get_current_societe(),
            "mode": "creation",
        }
    )


@app.post("/profils/nouveau", response_class=HTMLResponse)
async def profil_create(
    request: Request,
    nom: str = Form(...),
    prenom: str = Form(...),
    email: str = Form(...),
    telephone: str = Form(""),
    role: str = Form(...),
    actif: str = Form("oui"),
    certification_intitule: str = Form(""),
    certification_reference: str = Form(""),
    certification_date_obtention: str = Form(""),
    certification_date_validite: str = Form(""),
    certifie: str = Form("non"),
    certification_document: UploadFile | None = File(None),
):
    if role not in ROLES_ECHAFF:
        raise HTTPException(status_code=400, detail="Rôle invalide")

    profils = load_list_json(ECHAFF_PROFILS_FILE)

    certification_document_path = ""
    if certification_document and certification_document.filename:
        certification_document_path = save_upload_file(certification_document, DIPLOMES_DIR)

    profil = {
        "id": uuid4().hex,
        "nom": nom.strip(),
        "prenom": prenom.strip(),
        "email": email.strip(),
        "telephone": telephone.strip(),
        "societe": get_current_societe_name(),
        "role": role,
        "actif": actif == "oui",
        "signature_electronique": "",
        "certification": {
            "intitule": certification_intitule.strip(),
            "reference": certification_reference.strip(),
            "date_obtention": certification_date_obtention.strip(),
            "date_validite": certification_date_validite.strip(),
            "document": certification_document_path,
            "certifie": certifie == "oui",
        },
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
    }

    profils.append(profil)
    save_list_json(ECHAFF_PROFILS_FILE, profils)
    return RedirectResponse("/profils", status_code=303)


@app.get("/profils/{profil_id}/modifier", response_class=HTMLResponse)
def profil_edit_form(request: Request, profil_id: str):
    profil = get_profil_by_id(profil_id)
    if not profil:
        raise HTTPException(status_code=404, detail="Profil introuvable")

    return templates.TemplateResponse(
        request=request,
        name="profil_form.html",
        context={
            "request": request,
            "roles": ROLES_ECHAFF,
            "profil": profil,
            "societe": get_current_societe(),
            "mode": "edition",
        }
    )


@app.post("/profils/{profil_id}/modifier", response_class=HTMLResponse)
async def profil_update(
    request: Request,
    profil_id: str,
    nom: str = Form(...),
    prenom: str = Form(...),
    email: str = Form(...),
    telephone: str = Form(""),
    role: str = Form(...),
    actif: str = Form("oui"),
    certification_intitule: str = Form(""),
    certification_reference: str = Form(""),
    certification_date_obtention: str = Form(""),
    certification_date_validite: str = Form(""),
    certifie: str = Form("non"),
    certification_document: UploadFile | None = File(None),
):
    if role not in ROLES_ECHAFF:
        raise HTTPException(status_code=400, detail="Rôle invalide")

    profil = get_profil_by_id(profil_id)
    if not profil:
        raise HTTPException(status_code=404, detail="Profil introuvable")

    certification = profil.get("certification", {}) or {}
    certification_document_path = certification.get("document", "")

    if certification_document and certification_document.filename:
        certification_document_path = save_upload_file(certification_document, DIPLOMES_DIR)

    profil.update({
        "nom": nom.strip(),
        "prenom": prenom.strip(),
        "email": email.strip(),
        "telephone": telephone.strip(),
        "societe": get_current_societe_name(),
        "role": role,
        "actif": actif == "oui",
        "certification": {
            "intitule": certification_intitule.strip(),
            "reference": certification_reference.strip(),
            "date_obtention": certification_date_obtention.strip(),
            "date_validite": certification_date_validite.strip(),
            "document": certification_document_path,
            "certifie": certifie == "oui",
        },
    })

    save_profil(profil)
    return RedirectResponse("/profils", status_code=303)


@app.post("/profils/{profil_id}/supprimer")
def profil_delete(profil_id: str):
    deleted = delete_profil_by_id(profil_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Profil introuvable")

    return RedirectResponse("/profils", status_code=303)


@app.get("/api/profils")
def api_get_profils(role: str = ""):
    profils = load_list_json(ECHAFF_PROFILS_FILE)
    if role:
        profils = [p for p in profils if p.get("role") == role]
    return {"success": True, "profils": profils}


@app.post("/api/profils")
async def api_create_profil(request: Request):
    payload = await request.json()
    role = payload.get("role")

    if role not in ROLES_ECHAFF:
        raise HTTPException(status_code=400, detail="Rôle invalide")

    profils = load_list_json(ECHAFF_PROFILS_FILE)
    profil = {
        "id": uuid4().hex,
        "nom": payload.get("nom", "").strip(),
        "prenom": payload.get("prenom", "").strip(),
        "email": payload.get("email", "").strip(),
        "telephone": payload.get("telephone", "").strip(),
        "societe": get_current_societe_name(),
        "role": role,
        "actif": bool(payload.get("actif", True)),
        "signature_electronique": payload.get("signature_electronique", ""),
        "certification": {
            "intitule": payload.get("certification_intitule", "").strip(),
            "reference": payload.get("certification_reference", "").strip(),
            "date_obtention": payload.get("certification_date_obtention", "").strip(),
            "date_validite": payload.get("certification_date_validite", "").strip(),
            "document": payload.get("certification_document", "").strip(),
            "certifie": bool(payload.get("certifie", False)),
        },
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
    }
    profils.append(profil)
    save_list_json(ECHAFF_PROFILS_FILE, profils)
    return {"success": True, "profil": profil}


@app.put("/api/profils/{profil_id}")
async def api_update_profil(profil_id: str, request: Request):
    payload = await request.json()
    profil = get_profil_by_id(profil_id)
    if not profil:
        raise HTTPException(status_code=404, detail="Profil introuvable")

    role = payload.get("role", profil.get("role"))
    if role not in ROLES_ECHAFF:
        raise HTTPException(status_code=400, detail="Rôle invalide")

    certification = profil.get("certification", {}) or {}
    profil.update({
        "nom": payload.get("nom", profil.get("nom", "")).strip(),
        "prenom": payload.get("prenom", profil.get("prenom", "")).strip(),
        "email": payload.get("email", profil.get("email", "")).strip(),
        "telephone": payload.get("telephone", profil.get("telephone", "")).strip(),
        "societe": get_current_societe_name(),
        "role": role,
        "actif": bool(payload.get("actif", profil.get("actif", True))),
        "signature_electronique": payload.get("signature_electronique", profil.get("signature_electronique", "")),
        "certification": {
            "intitule": payload.get("certification_intitule", certification.get("intitule", "")).strip(),
            "reference": payload.get("certification_reference", certification.get("reference", "")).strip(),
            "date_obtention": payload.get("certification_date_obtention", certification.get("date_obtention", "")).strip(),
            "date_validite": payload.get("certification_date_validite", certification.get("date_validite", "")).strip(),
            "document": payload.get("certification_document", certification.get("document", "")).strip(),
            "certifie": bool(payload.get("certifie", certification.get("certifie", False))),
        },
    })
    save_profil(profil)
    return {"success": True, "profil": profil}


@app.delete("/api/profils/{profil_id}")
def api_delete_profil(profil_id: str):
    deleted = delete_profil_by_id(profil_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Profil introuvable")
    return {"success": True}


# ---------------------------------------------------------
# EXTENSION ECHAFF - CHANTIERS SOCIETE
# ---------------------------------------------------------

@app.get("/chantiers", response_class=HTMLResponse)
def chantiers_liste(request: Request):
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)

    for chantier in chantiers:
        chantier["pvs_reception"] = get_pvs_for_chantier(chantier.get("id", ""))
        chantier["nb_pvs"] = len(chantier["pvs_reception"])
        chantier["notifications"] = get_notifications_chantier(chantier.get("id", ""))

    return templates.TemplateResponse(
        request=request,
        name="chantiers_liste.html",
        context={
            "request": request,
            "chantiers": chantiers,
            "statuts": STATUTS_CHANTIER,
        }
    )


@app.get("/chantiers/nouveau", response_class=HTMLResponse)
def chantier_form(request: Request):
    societe = load_dict_json(ECHAFF_SOCIETE_FILE)
    return templates.TemplateResponse(
        request=request,
        name="chantier_form.html",
        context={"request": request, "statuts": STATUTS_CHANTIER, "societe": societe}
    )


@app.post("/chantiers/nouveau", response_class=HTMLResponse)
async def chantier_create(
    request: Request,
    nom: str = Form(...),
    adresse_complete: str = Form(""),
    batiment_zone_etage_secteur: str = Form(""),
    client_maitre_ouvrage: str = Form(""),
    date_debut: str = Form(""),
    date_fin_estimee: str = Form(""),
    date_fin_reelle: str = Form(""),
    statut: str = Form("brouillon"),
    societe_echafaudage_responsable: str = Form(""),
    societes_utilisatrices_autorisees: str = Form(""),
):
    if statut not in STATUTS_CHANTIER:
        raise HTTPException(status_code=400, detail="Statut chantier invalide")

    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)

    chantier = {
        "id": uuid4().hex,
        "nom": nom.strip(),
        "reference_interne": generate_next_chantier_reference(),
        "adresse_complete": adresse_complete.strip(),
        "batiment_zone_etage_secteur": batiment_zone_etage_secteur.strip(),
        "client_maitre_ouvrage": client_maitre_ouvrage.strip(),
        "date_debut": date_debut.strip(),
        "date_fin_estimee": date_fin_estimee.strip(),
        "date_fin_reelle": date_fin_reelle.strip(),
        "statut": statut,
        "societe_echafaudage_responsable": get_current_societe_name(),
        "societes_utilisatrices_autorisees": [
            s.strip() for s in societes_utilisatrices_autorisees.split(",") if s.strip()
        ],
        "documents_associes": [],
        "historique": [],
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
    }
    chantier = append_historique_chantier(chantier, "Création du chantier")

    chantiers.append(chantier)
    save_list_json(ECHAFF_CHANTIERS_FILE, chantiers)
    return RedirectResponse("/chantiers", status_code=303)


@app.get("/chantiers/{chantier_id}", response_class=HTMLResponse)
def chantier_detail(request: Request, chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    pvs_reception = get_pvs_for_chantier(chantier_id)

    return templates.TemplateResponse(
        request=request,
        name="chantier_detail.html",
        context={
            "request": request,
            "chantier": chantier,
            "pvs_reception": pvs_reception,
            "notifications": get_notifications_chantier(chantier_id),
            "qr_code_url": chantier.get("qr_code_url", ""),
            "qr_public_url": f"{APP_PUBLIC_URL}/qr/chantier/{chantier_id}",
            "statuts": STATUTS_CHANTIER,
        }
    )


@app.get("/chantiers/{chantier_id}/modifier", response_class=HTMLResponse)
def chantier_edit_form(request: Request, chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    return templates.TemplateResponse(
        request=request,
        name="chantier_form.html",
        context={
            "request": request,
            "chantier": chantier,
            "statuts": STATUTS_CHANTIER,
            "mode": "edition",
        }
    )


@app.post("/chantiers/{chantier_id}/modifier", response_class=HTMLResponse)
async def chantier_update(
    request: Request,
    chantier_id: str,
    nom: str = Form(...),
    reference_interne: str = Form(""),
    adresse_complete: str = Form(""),
    batiment_zone_etage_secteur: str = Form(""),
    client_maitre_ouvrage: str = Form(""),
    date_debut: str = Form(""),
    date_fin_estimee: str = Form(""),
    date_fin_reelle: str = Form(""),
    statut: str = Form("brouillon"),
    societe_echafaudage_responsable: str = Form(""),
    societes_utilisatrices_autorisees: str = Form(""),
):
    if statut not in STATUTS_CHANTIER:
        raise HTTPException(status_code=400, detail="Statut chantier invalide")

    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    chantier.update({
        "nom": nom.strip(),
        "reference_interne": reference_interne.strip(),
        "adresse_complete": adresse_complete.strip(),
        "batiment_zone_etage_secteur": batiment_zone_etage_secteur.strip(),
        "client_maitre_ouvrage": client_maitre_ouvrage.strip(),
        "date_debut": date_debut.strip(),
        "date_fin_estimee": date_fin_estimee.strip(),
        "date_fin_reelle": date_fin_reelle.strip(),
        "statut": statut,
        "societe_echafaudage_responsable": societe_echafaudage_responsable.strip(),
        "societes_utilisatrices_autorisees": [
            s.strip() for s in societes_utilisatrices_autorisees.split(",") if s.strip()
        ],
    })
    append_historique_chantier(chantier, "Modification du chantier")
    save_chantier(chantier)

    return RedirectResponse(f"/chantiers/{chantier_id}", status_code=303)


@app.post("/chantiers/{chantier_id}/qr")
def chantier_generate_qr(chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    qr_code_url = generate_qr_code_for_chantier(chantier_id)
    chantier["qr_code_url"] = qr_code_url
    chantier["qr_public_url"] = f"{APP_PUBLIC_URL}/qr/chantier/{chantier_id}"
    append_historique_chantier(chantier, "Génération du QR code chantier")
    save_chantier(chantier)

    return RedirectResponse(f"/chantiers/{chantier_id}", status_code=303)


@app.get("/qr/chantier/{chantier_id}", response_class=HTMLResponse)
def chantier_qr_public_page(request: Request, chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    return templates.TemplateResponse(
        request=request,
        name="qr_chantier.html",
        context={
            "request": request,
            "chantier": chantier,
            "pvs_reception": get_pvs_for_chantier(chantier_id),
            "notifications": get_notifications_chantier(chantier_id),
        }
    )


@app.get("/notifications")
def notifications():
    return {"success": True, "notifications": get_global_notifications()}


@app.post("/chantiers/{chantier_id}/archiver")
def chantier_archive(chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    chantier["statut"] = "archive"
    chantier["date_archivage"] = datetime.now().isoformat()
    append_historique_chantier(chantier, "Archivage du chantier")
    save_chantier(chantier)

    return RedirectResponse("/chantiers", status_code=303)


@app.post("/chantiers/{chantier_id}/supprimer")
def chantier_delete(chantier_id: str):
    deleted = delete_chantier_by_id(chantier_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    return RedirectResponse("/chantiers", status_code=303)


@app.get("/chantiers/{chantier_id}/pvs")
def chantier_pvs(chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    return {
        "success": True,
        "chantier": chantier,
        "pvs_reception": get_pvs_for_chantier(chantier_id),
        "notifications": get_notifications_chantier(chantier_id),
    }


@app.post("/chantiers/{chantier_id}/statut")
async def chantier_update_statut(chantier_id: str, request: Request):
    payload = await request.json()
    nouveau_statut = payload.get("statut", "")

    if nouveau_statut not in STATUTS_CHANTIER:
        raise HTTPException(status_code=400, detail="Statut chantier invalide")

    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    updated = None

    for chantier in chantiers:
        if chantier.get("id") == chantier_id:
            chantier["statut"] = nouveau_statut
            chantier["updated_at"] = datetime.now().isoformat()
            append_historique_chantier(chantier, f"Changement de statut : {nouveau_statut}")
            updated = chantier
            break

    if not updated:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    save_list_json(ECHAFF_CHANTIERS_FILE, chantiers)
    return {"success": True, "chantier": updated}


@app.get("/api/chantiers")
def api_get_chantiers():
    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)
    for chantier in chantiers:
        chantier["pvs_reception"] = get_pvs_for_chantier(chantier.get("id", ""))
        chantier["nb_pvs"] = len(chantier["pvs_reception"])
        chantier["notifications"] = get_notifications_chantier(chantier.get("id", ""))
    return {"success": True, "chantiers": chantiers}


@app.get("/api/chantiers/{chantier_id}")
def api_get_chantier_detail(chantier_id: str):
    chantier = get_chantier_by_id(chantier_id)
    if not chantier:
        raise HTTPException(status_code=404, detail="Chantier introuvable")

    return {
        "success": True,
        "chantier": chantier,
        "societe": get_current_societe(),
        "verificateur_defaut": get_default_verificateur_certifie(),
        "pvs_reception": get_pvs_for_chantier(chantier_id),
    }


@app.post("/api/chantiers")
async def api_create_chantier(request: Request):
    payload = await request.json()
    statut = payload.get("statut", "brouillon")

    if statut not in STATUTS_CHANTIER:
        raise HTTPException(status_code=400, detail="Statut chantier invalide")

    chantiers = load_list_json(ECHAFF_CHANTIERS_FILE)

    chantier = {
        "id": uuid4().hex,
        "nom": payload.get("nom", "").strip(),
        "reference_interne": generate_next_chantier_reference(),
        "adresse_complete": payload.get("adresse_complete", "").strip(),
        "batiment_zone_etage_secteur": payload.get("batiment_zone_etage_secteur", "").strip(),
        "client_maitre_ouvrage": payload.get("client_maitre_ouvrage", "").strip(),
        "date_debut": payload.get("date_debut", "").strip(),
        "date_fin_estimee": payload.get("date_fin_estimee", "").strip(),
        "date_fin_reelle": payload.get("date_fin_reelle", "").strip(),
        "statut": statut,
        "societe_echafaudage_responsable": get_current_societe_name(),
        "societes_utilisatrices_autorisees": payload.get("societes_utilisatrices_autorisees", []),
        "documents_associes": payload.get("documents_associes", []),
        "historique": [],
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat(),
    }
    chantier = append_historique_chantier(chantier, "Création du chantier")

    chantiers.append(chantier)
    save_list_json(ECHAFF_CHANTIERS_FILE, chantiers)
    return {"success": True, "chantier": chantier}


# =========================================================
# ROUTES - HOME / PV VERIFICATEUR
# =========================================================

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "request": request,
            "societe": get_current_societe(),
            "chantiers": load_list_json(ECHAFF_CHANTIERS_FILE),
            "profils": load_list_json(ECHAFF_PROFILS_FILE),
        }
    )


@app.post("/api/pv")
async def create_or_regenerate_pv(request: Request):
    try:
        raw_data = await request.json()
        data = prepare_pv_payload(raw_data)
        generated = regenerate_pv_files(data)

        emails = raw_data.get("emails_destinataires", "").strip()
        if not emails:
            emails = "; ".join([
                s.get("email", "").strip()
                for s in data.get("societes_utilisatrices", [])
                if s.get("email", "").strip()
            ])

        if emails and generated["pdf_path"]:
            pdf_url = f"{APP_PUBLIC_URL}/output/{generated['pdf_path'].name}"
            signature_url = f"{APP_PUBLIC_URL}/client-signature/{data['dossier_id']}"

            contenu = f"""Bonjour,

Veuillez trouver ci-dessous le PV de réception :

PDF :
{pdf_url}

Lien de vérification et signature :
{signature_url}

Cordialement,
OMNILUX
"""

            send_email(
                destinataires=emails,
                sujet=f"PV réception échafaudage n°{data['numero_pv']}",
                contenu=contenu
            )

        return {
            "success": True,
            "message": "PV généré avec succès",
            "dossier_id": data["dossier_id"],
            "numero_pv": data["numero_pv"],
            "excel_file": f"/output/{generated['xlsx_path'].name}",
            "pdf_file": f"/output/{generated['pdf_path'].name}" if generated["pdf_path"] else None,
            "json_file": f"/data/{data['dossier_id']}/state.json",
            "client_signature_url": f"/client-signature/{data['dossier_id']}",
        }

    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e)}


# =========================================================
# ROUTES - CLIENT SIGNATURE
# =========================================================

@app.get("/client-signature/{dossier_id}", response_class=HTMLResponse)
def client_signature_form(request: Request, dossier_id: str):
    paths = build_paths_for_dossier(dossier_id)

    if not paths["json_path"].exists():
        return templates.TemplateResponse(
            request=request,
            name="client_signature.html",
            context={
                "request": request,
                "dossier_id": dossier_id,
                "numero_pv": "",
                "chantier": "",
            }
        )

    dossier_data = load_json(paths["json_path"])

    return templates.TemplateResponse(
        request=request,
        name="client_signature.html",
        context={
            "request": request,
            "dossier_id": dossier_id,
            "numero_pv": dossier_data.get("numero_pv", ""),
            "chantier": dossier_data.get("chantier", ""),
        }
    )


@app.post("/client-signature/{dossier_id}")
async def client_signature_submit(dossier_id: str, request: Request):
    try:
        paths = build_paths_for_dossier(dossier_id)

        if not paths["json_path"].exists():
            raise HTTPException(status_code=404, detail="Dossier introuvable")

        payload = await request.json()
        dossier_data = load_json(paths["json_path"])
        dossier_data = apply_client_signature_payload(dossier_data, payload)
        generated = regenerate_pv_files(dossier_data)

        return {
            "success": True,
            "message": "Signature client enregistrée, PV régénéré",
            "dossier_id": dossier_id,
            "excel_file": f"/output/{generated['xlsx_path'].name}",
            "pdf_file": f"/output/{generated['pdf_path'].name}" if generated["pdf_path"] else None,
            "json_file": f"/data/{dossier_id}/state.json",
        }

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e)}


# =========================================================
# ROUTES - SIGNATURE D'UNE SOCIETE
# =========================================================

@app.post("/api/pv/{dossier_id}/societes/{societe_index}/sign")
async def sign_societe(dossier_id: str, societe_index: int, request: Request):
    try:
        payload = await request.json()
        signature_b64 = payload.get("signature_b64", "")

        paths = build_paths_for_dossier(dossier_id)

        if not paths["json_path"].exists():
            raise HTTPException(status_code=404, detail="Dossier introuvable")

        dossier_data = load_json(paths["json_path"])
        societes = dossier_data.get("societes_utilisatrices", [])

        if societe_index < 0 or societe_index >= len(societes):
            raise HTTPException(status_code=400, detail="Index société invalide")

        if not signature_b64:
            raise HTTPException(status_code=400, detail="Signature manquante")

        now = datetime.now()
        societes[societe_index]["signed"] = True
        societes[societe_index]["date_signature"] = now.strftime("%d/%m/%Y")
        societes[societe_index]["heure_signature"] = now.strftime("%H:%M")
        societes[societe_index]["signature_b64"] = signature_b64
        dossier_data["societes_utilisatrices"] = societes

        generated = regenerate_pv_files(dossier_data)

        return {
            "success": True,
            "message": "Signature société enregistrée, PV régénéré",
            "dossier_id": dossier_id,
            "excel_file": f"/output/{generated['xlsx_path'].name}",
            "pdf_file": f"/output/{generated['pdf_path'].name}" if generated["pdf_path"] else None,
            "json_file": f"/data/{dossier_id}/state.json",
        }

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e)}


# =========================================================
# ROUTES - VERIFICATEURS
# =========================================================

@app.get("/verificateurs/nouveau", response_class=HTMLResponse)
def form_verificateur(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="verificateur_form.html",
        context={"request": request}
    )


@app.post("/verificateurs/nouveau", response_class=HTMLResponse)
async def create_verificateur(
    request: Request,
    nom: str = Form(...),
    prenom: str = Form(...),
    email: str = Form(...),
    telephone: str = Form(""),
    numero_diplome: str = Form(...),
    date_obtention_diplome: str = Form(...),
    date_echeance_diplome: str = Form(...),
    carte_recto: UploadFile = File(...),
    carte_verso: UploadFile = File(...),
    diplome: UploadFile = File(...)
):
    fichier_carte_recto = save_upload_file(carte_recto, CARTES_DIR)
    fichier_carte_verso = save_upload_file(carte_verso, CARTES_DIR)
    fichier_diplome = save_upload_file(diplome, DIPLOMES_DIR)

    insert_verificateur(
        nom=nom.strip(),
        prenom=prenom.strip(),
        email=email.strip(),
        telephone=telephone.strip(),
        numero_diplome=numero_diplome.strip(),
        date_obtention_diplome=date_obtention_diplome.strip(),
        date_echeance_diplome=date_echeance_diplome.strip(),
        fichier_carte_recto=fichier_carte_recto,
        fichier_carte_verso=fichier_carte_verso,
        fichier_diplome=fichier_diplome
    )

    statut = get_diplome_status(date_echeance_diplome)

    return templates.TemplateResponse(
        request=request,
        name="verificateur_success.html",
        context={
            "request": request,
            "nom": nom,
            "prenom": prenom,
            "email": email,
            "telephone": telephone,
            "numero_diplome": numero_diplome,
            "date_obtention_diplome": date_obtention_diplome,
            "date_echeance_diplome": date_echeance_diplome,
            "statut_label": statut["label"],
            "statut_color": statut["color"]
        }
    )


# =========================================================
# ROUTES - ADMIN
# =========================================================

@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_form(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="admin_login.html",
        context={"request": request}
    )


@app.post("/admin/login", response_class=HTMLResponse)
async def admin_login(request: Request, password: str = Form(...)):
    if password == ADMIN_PASSWORD:
        request.session["is_admin"] = True
        return RedirectResponse("/admin/verificateurs", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="admin_login.html",
        context={"request": request, "error": "Mot de passe incorrect"}
    )


@app.get("/admin/logout")
def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/admin/login", status_code=303)


@app.get("/admin/verificateurs", response_class=HTMLResponse)
def liste_verificateurs(request: Request):
    if not request.session.get("is_admin"):
        return RedirectResponse("/admin/login", status_code=303)

    verificateurs_raw = get_all_verificateurs()
    verificateurs = []

    for v in verificateurs_raw:
        statut = get_diplome_status(v["date_echeance_diplome"])
        verificateurs.append({
            "id": v["id"],
            "nom": v["nom"],
            "prenom": v["prenom"],
            "email": v["email"],
            "telephone": v["telephone"],
            "numero_diplome": v["numero_diplome"],
            "date_obtention_diplome": v["date_obtention_diplome"],
            "date_echeance_diplome": v["date_echeance_diplome"],
            "fichier_carte_recto": v["fichier_carte_recto"],
            "fichier_carte_verso": v["fichier_carte_verso"],
            "fichier_diplome": v["fichier_diplome"],
            "actif": v["actif"],
            "statut_label": statut["label"],
            "statut_color": statut["color"]
        })

    return templates.TemplateResponse(
        request=request,
        name="verificateurs_liste.html",
        context={"request": request, "verificateurs": verificateurs}
    )


# =========================================================
# ROUTES - API VERIFICATEURS
# =========================================================

@app.get("/api/verificateurs")
def api_verificateurs(q: str = ""):
    verificateurs_raw = search_verificateurs(q)
    results = []

    for v in verificateurs_raw:
        statut = get_diplome_status(v["date_echeance_diplome"])

        results.append({
            "id": v["id"],
            "nom": v["nom"],
            "prenom": v["prenom"],
            "nom_complet": f'{v["nom"]} {v["prenom"]}',
            "email": v["email"],
            "telephone": v["telephone"],
            "numero_diplome": v["numero_diplome"],
            "date_obtention_diplome": v["date_obtention_diplome"],
            "date_echeance_diplome": v["date_echeance_diplome"],
            "fichier_diplome": f'/{v["fichier_diplome"]}' if v["fichier_diplome"] else "",
            "statut_label": statut["label"],
            "statut_color": statut["color"]
        })

    return JSONResponse(content=results)
